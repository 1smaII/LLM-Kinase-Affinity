from pathlib import Path
import argparse
import json
from dotenv import load_dotenv
from openpyxl import load_workbook
from providers import build_provider

ROOT = Path(__file__).parent
DATA = ROOT / "data"
RESULTS = ROOT / "results"

SEQ = "bindingdb target chain sequence"
IMG_PATH = "image path"
IMG_SLOT = "image slot"
SAMPLE = "sample id"

def norm(x):
    return " ".join(
        str(x or "").strip().casefold().replace("_", " ").split()
    )

def find_data():
    return sorted(
        DATA.glob("*/dataset.xlsx"),
        key=lambda p: p.parent.name.casefold(),
    )

def process_sheet(rows, seqs, catalog, images):
    for h, row in enumerate(rows):
        cols = [norm(x) for x in row]

        if SEQ not in cols and IMG_PATH not in cols:
            continue

        seq_i = cols.index(SEQ) if SEQ in cols else None
        path_i = cols.index(IMG_PATH) if IMG_PATH in cols else None
        slot_i = cols.index(IMG_SLOT) if IMG_SLOT in cols else None
        id_i = cols.index(SAMPLE) if SAMPLE in cols else None

        if seq_i is not None:
            row[seq_i] = "Sequence_ID"

        for r in rows[h + 1:]:
            if seq_i is not None and seq_i < len(r):
                seq = "".join(str(r[seq_i] or "").split()).upper()

                if seq:
                    if seq not in seqs:
                        sid = f"S{len(seqs) + 1}"
                        seqs[seq] = sid
                        catalog.append([sid, seq])
                    r[seq_i] = seqs[seq]

            if path_i is None or path_i >= len(r):
                continue

            if r[path_i] in (None, ""):
                continue
            path = str(r[path_i]).strip()

            slot = ""
            if slot_i is not None and slot_i < len(r):
                slot = str(r[slot_i] or "").strip()

            sample = ""
            if id_i is not None and id_i < len(r):
                sample = str(r[id_i] or "").strip()

            pair = (slot, sample)

            if pair != ("", ""):
                group = images.setdefault(path, [])
                if pair not in group:
                    group.append(pair)

        drop = {
            i for i in (path_i, slot_i)
            if i is not None
        }

        if drop:
            for i in range(h, len(rows)):
                rows[i] = [
                    x for j, x in enumerate(rows[i])
                    if j not in drop
                ]
        break
    return rows

def prepare_input(xlsx):
    wb = load_workbook(
        xlsx,
        data_only=True,
        read_only=True,
    )

    seqs = {}
    catalog = []
    images = {}
    sheets = []
    for ws in wb.worksheets:
        rows = [
            list(r)
            for r in ws.iter_rows(values_only=True)
        ]

        rows = process_sheet(rows, seqs, catalog, images)
        sheets.append({
            "sheet_name": ws.title,
            "rows": rows,
        })
    wb.close()
    data = { "file_name": xlsx.name,"sheets": sheets,}
    if catalog:
        data["sequence_catalog"] = {
            "columns": [
                "Sequence_ID",
                "BindingDB Target Chain Sequence",
            ],
            "rows": catalog,
        }

    imgs = [
        {
            "path": path,
            "slots": slots,
        }
        for path, slots in images.items()
    ]

    text = json.dumps(data,ensure_ascii=False,default=str,separators=(",", ":"), )
    return text, imgs

def save(mode, model, xlsx, text):
    folder = RESULTS / mode / model.name
    folder.mkdir(
        parents=True,
        exist_ok=True,
    )
    out = folder / f"{xlsx.parent.name}.csv"

    out.write_text(
        text.strip() + "\n",
        encoding="utf-8-sig",
    )
    print(f"已保存：{out}")

def main():
    load_dotenv(ROOT / ".env")
    cfg = json.loads(
        (ROOT / "benchmark_config.json").read_text(
            encoding="utf-8"
        )
    )

    parser = argparse.ArgumentParser()
    parser.add_argument(
        "mode",
        choices=["text", "multimodal"],
    )

    parser.add_argument("model", nargs="?",default="all", choices=["qwen", "gpt", "gemini", "all"], )
    args = parser.parse_args()

    if args.model == "all":
        names = ["qwen", "gpt", "gemini"]
    else:
        names = [args.model]

    models = [build_provider(name, args.mode,cfg[name],cfg["max_tokens"], )
        for name in names
    ]

    prompt = (ROOT / cfg["prompts"][args.mode]).read_text(encoding="utf-8" ).strip()
    files = find_data()

    if not files:
        raise FileNotFoundError(
            "data 下没有找到 */dataset.xlsx"
        )

    for xlsx in files:
        print(f"\n数据集：{xlsx.parent.name}")
        text, images = prepare_input(xlsx)

        if args.mode == "multimodal":
            print(f"图片数：{len(images)}")

        for model in models:
            print(
                f"{args.mode} | "
                f"{model.name} | "
                f"{model.model}"
            )
            result = model.generate(prompt, text,images,xlsx, )
            save(args.mode,model, xlsx,result, )
    print("\n全部完成。")

if __name__ == "__main__":
    main()
